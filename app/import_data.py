import os

from flask_script import Command
import openpyxl
from datetime import datetime
from app.config import BASE_DIR
from app import db
from app.models import Car_data, Info, Car_data_new


class ImportData(Command):
    def run(self):
        print('导入数据开始')
        dir = os.path.join(BASE_DIR, 'c_excel')
        for file_name in os.listdir(dir):
            file_path = os.path.join(dir, file_name)
            self.save_to_mysql(file_path)
        print('导入数据完成')

    def save_to_mysql(self, file_path):

        fileds = ['metro_id', 'car_id', 'wheel_id',
                  'QR', 'Sh', 'Sd', 'ddM', 'dAR', 'WD', 'AR', 'date']

        wb = openpyxl.load_workbook(file_path)
        index = wb.sheetnames
        ws = wb[index[0]]
        for row in ws.iter_rows(min_row=2):
            data = [cell.value for cell in row]
            dict_val = dict(zip(fileds, data))
            data = Car_data(**dict_val)
            wheel_data = Car_data(metro_id=dict_val['metro_id'], car_id=dict_val['car_id'],
                                  wheel_id=dict_val['wheel_id'], date=dict_val['date'])
            newdata = Car_data_new(**dict_val)
            wheel_newdata = Car_data_new(metro_id=dict_val['metro_id'], car_id=dict_val['car_id'],
                                         wheel_id=dict_val['wheel_id'])
            self.save_data(data, wheel_data)
            self.save_new_data(newdata, wheel_newdata)

        def average(col):
            avg = 0
            for i in range(2, 50):
                avg += ws[f'{col}{i}'].value / 48
            return avg

        dict_val2 = {'metro_id': ws['A2'].value, 'date': ws['K2'].value, 'QR_avg':average('D'),
                     'Sh_avg':average('E'),'Sd_avg': average('F'), 'WD_avg': average('I'),'mile':ws['L2'].value}
        info = Info(**dict_val2)
        info_data = Info(metro_id=dict_val2['metro_id'], date=dict_val2['date'],
                         Sd_avg=dict_val2['Sd_avg'], WD_avg=dict_val2['WD_avg'],
                         QR_avg=dict_val2['QR_avg'],Sh_avg=dict_val2['Sh_avg'],mile=dict_val2['mile'])
        self.save_info(info, info_data)

        wb.close()

    def save_data(self, data, wheel_data):
        try:
            data1 = Car_data.query.filter_by(metro_id=wheel_data.metro_id, car_id=wheel_data.car_id,
                                             wheel_id=wheel_data.wheel_id, date=wheel_data.date).first()
            if not data1:
                db.session.merge(data)
                db.session.commit()
        except:
            db.session.rollback()

    def save_info(self, info, info_data):
        try:
            data = Info.query.filter_by(metro_id=info_data.metro_id, date=info_data.date).first()
            if not data:
                db.session.add(info)
                db.session.commit()
        except:
            db.session.rollback()

    def save_new_data(self, newdata, wheel_newdata):
        try:
            data1 = Car_data_new.query.filter_by(metro_id=wheel_newdata.metro_id, car_id=wheel_newdata.car_id,
                                                 wheel_id=wheel_newdata.wheel_id).all()
            for i in data1:
                a = i.date.strftime("%Y-%m-%d")
                day = datetime.strptime(a, "%Y-%m-%d")
                diff_day = (day - wheel_newdata.date).days

                if diff_day <= 0:

                    db.session.query(Car_data_new).filter_by(metro_id=wheel_newdata.metro_id).delete()
            data2 = Car_data_new.query.filter_by(metro_id=wheel_newdata.metro_id, car_id=wheel_newdata.car_id,
                                                 wheel_id=wheel_newdata.wheel_id).first()
            if not data2:
                db.session.merge(newdata)
                db.session.commit()
        except:
            db.session.rollback()


if __name__ == '__main__':
    print(BASE_DIR)
